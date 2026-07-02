"""
Formularios SRI Ecuador + Consulta RUC
- Consulta RUC en tiempo real al SRI
- Formulario 104: Declaración mensual de IVA
- ATS: Anexo Transaccional Simplificado (mensual)
- RDEP: Relación de dependencia (anual)
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from database import query, query_one, execute, insert
from auth import get_current_user
import httpx, re
from typing import Optional
from datetime import date
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/api/sri", tags=["SRI Formularios"])


# ══════════════════════════════════════════════════════════════
#  CONSULTA RUC / CÉDULA EN EL SRI
# ══════════════════════════════════════════════════════════════

@router.get("/consulta-ruc/{ruc}")
async def consultar_ruc(ruc: str, u=Depends(get_current_user)):
    """
    Consulta los datos de un RUC o cédula en el SRI Ecuador.
    Usa el portal público del SRI (consulta de contribuyentes).
    """
    ruc = ruc.strip()
    if not re.match(r'^\d{10,13}$', ruc):
        raise HTTPException(400, "RUC o cédula inválido — debe tener 10 o 13 dígitos")

    # Primero buscar en la base de datos local (clientes/proveedores existentes)
    if len(ruc) == 13:
        local = query_one("""
            SELECT identificacion as ruc, razon_social, direccion, email, telefono,
                   tipo_contribuyente, obligado_contabilidad, 'cliente' as origen
            FROM ven_clientes WHERE identificacion = %s
            UNION ALL
            SELECT identificacion, razon_social, direccion, email, telefono,
                   tipo_contribuyente, obligado_contabilidad, 'proveedor' as origen
            FROM com_proveedores WHERE identificacion = %s
            LIMIT 1
        """, (ruc, ruc))
        if local:
            return {**dict(local), "fuente": "sistema"}

    # Consultar al SRI
    try:
        async with httpx.AsyncClient(timeout=10, verify=False) as client:
            # Portal público SRI — consulta de RUC
            url = f"https://srienlinea.sri.gob.ec/sri-en-linea/SriRucWeb/ConsultaRuc/Consultas/consultaRuc"
            resp = await client.get(url, params={"ruc": ruc},
                headers={"Accept": "application/json",
                         "User-Agent": "Mozilla/5.0"})

            if resp.status_code == 200:
                try:
                    data = resp.json()
                    contribuyente = data.get("contribuyente") or data
                    return {
                        "ruc":                    ruc,
                        "razon_social":           contribuyente.get("razonSocial", ""),
                        "nombre_comercial":       contribuyente.get("nombreFantasia", ""),
                        "tipo_contribuyente":     contribuyente.get("tipoContribuyente", ""),
                        "obligado_contabilidad":  contribuyente.get("obligadoLlevarContabilidad","NO") == "SI",
                        "estado":                 contribuyente.get("estado", ""),
                        "actividad_principal":    contribuyente.get("actividadPrincipal", ""),
                        "direccion":              contribuyente.get("direccionMatriz", ""),
                        "fuente": "sri"
                    }
                except Exception:
                    pass

            # Fallback: scraping del portal público
            resp2 = await client.get(
                f"https://srienlinea.sri.gob.ec/facturacion-internet/pages/consultas/consulta-ruc.jsf",
                params={"ruc": ruc})
            # Extraer datos básicos del HTML
            html = resp2.text
            razon = re.search(r'Razón Social[:\s]*</[^>]+>\s*<[^>]+>([^<]+)', html)
            return {
                "ruc": ruc,
                "razon_social": razon.group(1).strip() if razon else "",
                "fuente": "sri_html",
                "raw": len(html)
            }

    except httpx.TimeoutException:
        raise HTTPException(504, "El SRI no respondió. Intente de nuevo en unos segundos.")
    except Exception as e:
        # Si falla la consulta al SRI, devolver vacío para que el usuario llene manual
        return {
            "ruc": ruc,
            "razon_social": "",
            "fuente": "sin_datos",
            "mensaje": "No se pudo consultar el SRI. Ingresa los datos manualmente.",
            "error": str(e)[:100]
        }


# ══════════════════════════════════════════════════════════════
#  FORMULARIO 104 — IVA MENSUAL
# ══════════════════════════════════════════════════════════════

@router.get("/formulario104")
def calcular_104(periodo: str, u=Depends(get_current_user)):
    """
    Calcula los valores del Formulario 104 para un período (YYYY-MM).
    Basado en facturas de ventas y compras del período.
    """
    anio, mes = int(periodo[:4]), int(periodo[5:7])

    # ── VENTAS ────────────────────────────────────────────────
    ventas = query_one("""
        SELECT
            COALESCE(SUM(subtotal_0),0)   as base_0,
            COALESCE(SUM(subtotal_iva),0) as base_gravada,
            COALESCE(SUM(iva),0)          as iva_cobrado,
            COALESCE(SUM(total),0)        as total_ventas,
            COUNT(*) FILTER (WHERE estado != 'ANULADA') as num_facturas
        FROM ven_facturas
        WHERE EXTRACT(YEAR FROM fecha_emision) = %s
          AND EXTRACT(MONTH FROM fecha_emision) = %s
          AND estado NOT IN ('ANULADA','BORRADOR','HISTORICO')
    """, (anio, mes)) or {}

    # ── COMPRAS (crédito tributario) ──────────────────────────
    compras = query_one("""
        SELECT
            COALESCE(SUM(subtotal_0),0) as base_0,
            COALESCE(SUM(subtotal_iva),0) as base_gravada,
            COALESCE(SUM(iva),0) as iva_pagado,
            COUNT(*) as num_compras
        FROM com_compras
        WHERE EXTRACT(YEAR FROM fecha) = %s
          AND EXTRACT(MONTH FROM fecha) = %s
          AND estado NOT IN ('ANULADA','HISTORICO')
    """, (anio, mes)) or {}

    # ── RETENCIONES RECIBIDAS ─────────────────────────────────
    ret_recibidas = query_one("""
        SELECT COALESCE(SUM(valor_retenido),0) as total
        FROM ven_retenciones
        WHERE EXTRACT(YEAR FROM fecha_emision) = %s
          AND EXTRACT(MONTH FROM fecha_emision) = %s
          AND tipo_impuesto = 'IVA'
    """, (anio, mes)) or {"total": 0}

    # ── CÁLCULO DEL FORMULARIO ────────────────────────────────
    base_0          = float(ventas.get("base_0", 0))
    base_gravada    = float(ventas.get("base_gravada", 0))
    iva_cobrado     = float(ventas.get("iva_cobrado", 0))
    iva_pagado      = float(compras.get("iva_pagado", 0))
    ret_iva         = float(ret_recibidas.get("total", 0))
    credito_trib    = iva_pagado
    iva_a_pagar     = max(0, iva_cobrado - credito_trib - ret_iva)
    credito_favor   = max(0, credito_trib + ret_iva - iva_cobrado)

    return {
        "periodo":           periodo,
        "empresa":           query_one("SELECT ruc, razon_social FROM sys_empresas WHERE activa=true LIMIT 1"),
        # Ventas
        "ventas_base_0":     round(base_0, 2),
        "ventas_base_iva":   round(base_gravada, 2),
        "ventas_iva":        round(iva_cobrado, 2),
        "ventas_total":      round(float(ventas.get("total_ventas", 0)), 2),
        "num_facturas":      int(ventas.get("num_facturas", 0)),
        # Compras
        "compras_base_0":    round(float(compras.get("base_0", 0)), 2),
        "compras_base_iva":  round(float(compras.get("base_gravada", 0)), 2),
        "compras_iva":       round(iva_pagado, 2),
        "num_compras":       int(compras.get("num_compras", 0)),
        # Resultado
        "credito_tributario": round(credito_trib, 2),
        "retenciones_recibidas": round(ret_iva, 2),
        "iva_a_pagar":       round(iva_a_pagar, 2),
        "credito_a_favor":   round(credito_favor, 2),
        # Campos formulario SRI (casillas)
        "casilla_401": round(base_0, 2),        # Ventas netas tarifa 0%
        "casilla_411": round(base_gravada, 2),   # Ventas netas tarifa 15%
        "casilla_421": round(iva_cobrado, 2),    # IVA cobrado
        "casilla_500": round(float(compras.get("base_0",0)), 2),     # Compras 0%
        "casilla_510": round(float(compras.get("base_gravada",0)), 2), # Compras 15%
        "casilla_520": round(iva_pagado, 2),     # Crédito tributario
        "casilla_601": round(ret_iva, 2),        # Retenciones IVA que le hicieron
        "casilla_699": round(iva_a_pagar, 2),    # IVA a pagar
    }


@router.get("/formulario104/excel")
def exportar_104_excel(periodo: str, u=Depends(get_current_user)):
    """Exporta el Formulario 104 en Excel con formato."""
    data = calcular_104(periodo, u)
    empresa = data.get("empresa") or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"F104 {periodo}"

    # Estilos
    hdr = Font(bold=True, color="FFFFFF", size=11)
    bg_blue   = PatternFill("solid", fgColor="1D4ED8")
    bg_dark   = PatternFill("solid", fgColor="0F172A")
    bg_green  = PatternFill("solid", fgColor="064E3B")
    bg_red    = PatternFill("solid", fgColor="7F1D1D")
    bg_gray   = PatternFill("solid", fgColor="1E293B")
    center    = Alignment(horizontal="center", vertical="center")
    right_al  = Alignment(horizontal="right", vertical="center")
    thin      = Side(style="thin", color="334155")
    bord      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, val, bold=False, bg=None, color="E2E8F0", align=None, num=False):
        c = ws.cell(row, col, val)
        if bold: c.font = Font(bold=True, color=color, size=11)
        else:    c.font = Font(color=color, size=10)
        if bg:   c.fill = bg
        c.alignment = align or Alignment(vertical="center", wrap_text=True, indent=1)
        c.border = bord
        if num and val is not None:
            c.number_format = '#,##0.00'
        return c

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20

    # Título
    ws.merge_cells("A1:C1")
    t = ws.cell(1, 1, "FORMULARIO 104 — DECLARACIÓN DE IVA")
    t.font = Font(bold=True, color="FFFFFF", size=14)
    t.fill = bg_dark; t.alignment = center

    ws.merge_cells("A2:C2")
    e = ws.cell(2, 1, f"{empresa.get('razon_social','')} | RUC: {empresa.get('ruc','')} | Período: {periodo}")
    e.font = Font(color="94A3B8", size=10)
    e.fill = bg_dark; e.alignment = center

    row = 4

    def seccion(titulo, bg):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row, 1, titulo)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = bg; c.alignment = center
        ws.row_dimensions[row].height = 22
        row += 1

    def fila(casilla, desc, valor):
        nonlocal row
        cell(row, 1, casilla, bold=True, bg=bg_gray, color="93C5FD", align=center)
        cell(row, 2, desc, bg=bg_gray, color="E2E8F0")
        cell(row, 3, valor, bg=bg_gray, color="6EE7B7", align=right_al, num=True)
        ws.row_dimensions[row].height = 18
        row += 1

    # VENTAS
    seccion("▶ VENTAS / INGRESOS", bg_blue)
    fila("401", "Ventas netas tarifa 0% de IVA", data["casilla_401"])
    fila("411", "Ventas netas tarifa 15% de IVA", data["casilla_411"])
    fila("421", "IVA cobrado en ventas (15%)", data["casilla_421"])
    row += 1

    # COMPRAS
    seccion("▶ COMPRAS / CRÉDITO TRIBUTARIO", bg_blue)
    fila("500", "Compras netas tarifa 0%", data["casilla_500"])
    fila("510", "Compras netas tarifa 15%", data["casilla_510"])
    fila("520", "IVA en compras (Crédito tributario)", data["casilla_520"])
    fila("601", "Retenciones de IVA que le realizaron", data["casilla_601"])
    row += 1

    # RESULTADO
    seccion("▶ RESULTADO — IVA A PAGAR / CRÉDITO A FAVOR", bg_red if data["iva_a_pagar"] > 0 else bg_green)
    fila("699", "IVA A PAGAR (si ventas > compras)", data["casilla_699"])
    fila("—",  "Crédito a favor (si compras > ventas)", data["credito_a_favor"])
    row += 2

    # Nota
    ws.merge_cells(f"A{row}:C{row}")
    n = ws.cell(row, 1, "⚠️ Este documento es referencial. Verifique con su contador antes de declarar al SRI.")
    n.font = Font(color="F59E0B", size=9, italic=True)
    n.fill = PatternFill("solid", fgColor="1C1917")
    n.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Formulario104_{periodo}.xlsx"})


# ══════════════════════════════════════════════════════════════
#  ATS — ANEXO TRANSACCIONAL SIMPLIFICADO
# ══════════════════════════════════════════════════════════════

@router.get("/ats")
def calcular_ats(periodo: str, u=Depends(get_current_user)):
    """
    Genera los datos del ATS para el período.
    El ATS incluye detalle de cada factura de venta y compra.
    """
    anio, mes = int(periodo[:4]), int(periodo[5:7])

    ventas = query("""
        SELECT f.numero_factura, f.fecha_emision,
               cl.tipo_identificacion, cl.identificacion,
               cl.razon_social as nombre,
               f.subtotal_0, f.subtotal_iva, f.iva, f.total,
               f.forma_pago
        FROM ven_facturas f
        JOIN ven_clientes cl ON cl.id = f.cliente_id
        WHERE EXTRACT(YEAR FROM f.fecha_emision) = %s
          AND EXTRACT(MONTH FROM f.fecha_emision) = %s
          AND f.estado NOT IN ('ANULADA','BORRADOR','HISTORICO')
        ORDER BY f.fecha_emision, f.numero_factura
    """, (anio, mes))

    compras = query("""
        SELECT c.numero_factura_prov, c.fecha,
               p.tipo_identificacion, p.identificacion,
               p.razon_social as nombre,
               c.subtotal_0, c.subtotal_iva, c.iva, c.total
        FROM com_compras c
        JOIN com_proveedores p ON p.id = c.proveedor_id
        WHERE EXTRACT(YEAR FROM c.fecha) = %s
          AND EXTRACT(MONTH FROM c.fecha) = %s
          AND c.estado NOT IN ('ANULADA','HISTORICO')
        ORDER BY c.fecha
    """, (anio, mes))

    return {
        "periodo": periodo,
        "empresa": query_one("SELECT ruc, razon_social FROM sys_empresas WHERE activa=true LIMIT 1"),
        "ventas":  [dict(v) for v in ventas],
        "compras": [dict(c) for c in compras],
        "resumen": {
            "total_ventas":  len(ventas),
            "total_compras": len(compras),
            "base_iva_ventas":  sum(float(v["subtotal_iva"] or 0) for v in ventas),
            "iva_ventas":       sum(float(v["iva"] or 0) for v in ventas),
            "base_iva_compras": sum(float(c["subtotal_iva"] or 0) for c in compras),
            "iva_compras":      sum(float(c["iva"] or 0) for c in compras),
        }
    }


@router.get("/ats/excel")
def exportar_ats_excel(periodo: str, u=Depends(get_current_user)):
    """Exporta el ATS en Excel (2 hojas: Ventas y Compras)."""
    data = calcular_ats(periodo, u)
    empresa = data.get("empresa") or {}

    wb = openpyxl.Workbook()

    def crear_hoja(ws_nombre, registros, cols_headers, cols_keys):
        ws = wb.create_sheet(ws_nombre)
        bg_h = PatternFill("solid", fgColor="1D4ED8")
        bg_r = PatternFill("solid", fgColor="0F172A")
        bg_a = PatternFill("solid", fgColor="111827")
        thin = Side(style="thin", color="1F2937")
        bord = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Título
        ws.merge_cells(f"A1:{chr(64+len(cols_headers))}1")
        t = ws.cell(1, 1, f"ATS {ws_nombre} — {empresa.get('razon_social','')} — {periodo}")
        t.font = Font(bold=True, color="FFFFFF", size=12)
        t.fill = PatternFill("solid", fgColor="0F172A")
        t.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        # Cabeceras
        for i, h in enumerate(cols_headers, 1):
            c = ws.cell(2, i, h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = bg_h
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bord

        # Datos
        for ri, reg in enumerate(registros, 3):
            bg = bg_r if ri % 2 == 1 else bg_a
            for ci, key in enumerate(cols_keys, 1):
                val = reg.get(key, "")
                if isinstance(val, (int, float)) or (isinstance(val, str) and val.replace(".","").isdigit()):
                    try: val = float(val)
                    except: pass
                c = ws.cell(ri, ci, val)
                c.font = Font(color="E2E8F0", size=10)
                c.fill = bg; c.border = bord
                if isinstance(val, float):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")

        # Anchos
        for i in range(1, len(cols_headers)+1):
            ws.column_dimensions[chr(64+i)].width = 20
        return ws

    # Hoja ventas
    crear_hoja("Ventas", data["ventas"],
        ["N° Factura","Fecha","Tipo ID","Identificación","Cliente","Base 0%","Base IVA","IVA","Total"],
        ["numero_factura","fecha_emision","tipo_identificacion","identificacion","nombre",
         "subtotal_0","subtotal_iva","iva","total"])

    # Hoja compras
    crear_hoja("Compras", data["compras"],
        ["N° Factura Prov.","Fecha","Tipo ID","RUC/Cédula","Proveedor","Base 0%","Base IVA","IVA","Total"],
        ["numero_factura_prov","fecha","tipo_identificacion","identificacion","nombre",
         "subtotal_0","subtotal_iva","iva","total"])

    # Quitar hoja vacía inicial
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ATS_{periodo}.xlsx"})


# ══════════════════════════════════════════════════════════════
#  RDEP — RELACIÓN DE DEPENDENCIA (anual)
# ══════════════════════════════════════════════════════════════

@router.get("/rdep")
def calcular_rdep(anio: int, u=Depends(get_current_user)):
    """
    Calcula el RDEP para el año indicado.
    Muestra el resumen anual de ingresos por empleado para la declaración.
    """
    empleados = query("""
        SELECT e.cedula, e.nombres, e.apellidos, e.email,
               e.cargo, e.fecha_ingreso, e.fecha_salida, e.region,
               COALESCE(SUM(r.total_ingresos),0)       as total_ingresos,
               COALESCE(SUM(r.aporte_iess_personal),0) as aporte_iess,
               COALESCE(SUM(r.bonificaciones),0)       as bonificaciones,
               COALESCE(SUM(r.neto_a_pagar),0)         as total_neto,
               COUNT(r.id)                             as num_roles
        FROM nom_empleados e
        LEFT JOIN nom_roles_pago r ON r.empleado_id = e.id
            AND EXTRACT(YEAR FROM r.created_at) = %s
            AND r.estado = 'APROBADO'
        WHERE e.activo = true OR (e.fecha_salida IS NOT NULL
              AND EXTRACT(YEAR FROM e.fecha_salida) = %s)
        GROUP BY e.id, e.cedula, e.nombres, e.apellidos, e.email,
                 e.cargo, e.fecha_ingreso, e.fecha_salida, e.region
        ORDER BY e.apellidos, e.nombres
    """, (anio, anio))

    empresa = query_one("SELECT * FROM sys_empresas WHERE activa=true LIMIT 1") or {}

    return {
        "anio":      anio,
        "empresa":   empresa,
        "empleados": [dict(e) for e in empleados],
        "resumen": {
            "total_empleados": len(empleados),
            "total_ingresos":  sum(float(e["total_ingresos"]) for e in empleados),
            "total_aporte_iess": sum(float(e["aporte_iess"]) for e in empleados),
        }
    }


@router.get("/rdep/excel")
def exportar_rdep_excel(anio: int, u=Depends(get_current_user)):
    data = calcular_rdep(anio, u)
    empresa = data.get("empresa") or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"RDEP {anio}"

    bg_dark = PatternFill("solid", fgColor="0F172A")
    bg_blue = PatternFill("solid", fgColor="1D4ED8")
    bg_row1 = PatternFill("solid", fgColor="111827")
    bg_row2 = PatternFill("solid", fgColor="0F172A")
    thin    = Side(style="thin", color="1F2937")
    bord    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    t = ws.cell(1, 1, f"RDEP {anio} — {empresa.get('razon_social','')} | RUC: {empresa.get('ruc','')}")
    t.font = Font(bold=True, color="FFFFFF", size=13)
    t.fill = bg_dark; t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    headers = ["Cédula","Nombres","Apellidos","Cargo","Fecha Ingreso","Fecha Salida",
               "Total Ingresos","Aporte IESS Personal","Bonificaciones","Neto Pagado"]
    keys    = ["cedula","nombres","apellidos","cargo","fecha_ingreso","fecha_salida",
               "total_ingresos","aporte_iess","bonificaciones","total_neto"]

    for i, h in enumerate(headers, 1):
        c = ws.cell(2, i, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = bg_blue; c.border = bord
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64+i)].width = 18

    for ri, emp in enumerate(data["empleados"], 3):
        bg = bg_row1 if ri % 2 == 1 else bg_row2
        for ci, key in enumerate(keys, 1):
            val = emp.get(key, "")
            if val and isinstance(val, str) and "T" in str(val): val = str(val)[:10]
            c = ws.cell(ri, ci, val)
            c.font = Font(color="E2E8F0", size=10)
            c.fill = bg; c.border = bord
            if ci >= 7 and val:
                try: c.value = float(val); c.number_format = '#,##0.00'
                except: pass
                c.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=RDEP_{anio}.xlsx"})
