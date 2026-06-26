"""
Módulo de Reportes — NEXUS POS
Reportes con filtros flexibles y exportación a Excel/PDF.
"""
import io
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response
from database import query, query_one
from auth import get_current_user
from typing import Optional

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])


# ══════════════════════════════════════════════════════════════
#  1. VENTAS
# ══════════════════════════════════════════════════════════════

@router.get("/ventas")
def reporte_ventas(
    fecha_ini: str, fecha_fin: str,
    cliente_id:  Optional[int] = None,
    cliente:     Optional[str] = None,
    vendedor_id: Optional[int] = None,
    sucursal_id: Optional[int] = None,
    u=Depends(get_current_user)
):
    conds  = ["f.estado='EMITIDA'", "f.fecha_emision BETWEEN %s AND %s"]
    params = [fecha_ini, fecha_fin]
    if cliente_id:  conds.append("f.cliente_id=%s");  params.append(cliente_id)
    if cliente:
        conds.append("(c.razon_social ILIKE %s OR c.identificacion ILIKE %s)")
        params += [f"%{cliente}%", f"%{cliente}%"]
    if vendedor_id: conds.append("f.vendedor_id=%s"); params.append(vendedor_id)
    if sucursal_id: conds.append("f.sucursal_id=%s"); params.append(sucursal_id)
    where = "WHERE " + " AND ".join(conds)

    resumen = query_one(f"""
        SELECT COUNT(*) as facturas,
               COALESCE(SUM(f.subtotal_0),0)  as subtotal_0,
               COALESCE(SUM(f.subtotal_iva),0) as subtotal_iva,
               COALESCE(SUM(f.iva),0)          as iva,
               COALESCE(SUM(f.total),0)        as total
        FROM ven_facturas f
        JOIN ven_clientes c ON c.id=f.cliente_id
        {where}
    """, params)

    detalle = query(f"""
        SELECT f.id, f.numero_factura, f.fecha_emision,
               c.razon_social as cliente, c.identificacion as ruc,
               COALESCE(v.nombre,'') as vendedor,
               s.nombre as sucursal,
               f.subtotal_0, f.subtotal_iva, f.iva, f.total,
               f.descuento_global_pct
        FROM ven_facturas f
        JOIN ven_clientes c        ON c.id=f.cliente_id
        LEFT JOIN ven_vendedores v ON v.id=f.vendedor_id
        LEFT JOIN sys_sucursales s ON s.id=f.sucursal_id
        {where}
        ORDER BY f.fecha_emision, f.id
    """, params)

    por_dia = query(f"""
        SELECT f.fecha_emision::date as fecha,
               COUNT(*) as facturas, SUM(f.total) as total
        FROM ven_facturas f
        JOIN ven_clientes c ON c.id=f.cliente_id
        {where}
        GROUP BY f.fecha_emision::date ORDER BY fecha
    """, params)

    por_vendedor = query(f"""
        SELECT COALESCE(v.nombre,'Sin vendedor') as vendedor,
               COUNT(*) as facturas, SUM(f.total) as total
        FROM ven_facturas f
        JOIN ven_clientes c ON c.id=f.cliente_id
        LEFT JOIN ven_vendedores v ON v.id=f.vendedor_id
        {where}
        GROUP BY v.nombre ORDER BY total DESC
    """, params)

    por_forma_pago = query(f"""
        SELECT p.forma_pago, COUNT(*) as operaciones,
               SUM(p.monto) as total
        FROM ven_pagos p
        JOIN ven_facturas f ON f.id=p.factura_id
        JOIN ven_clientes c ON c.id=f.cliente_id
        {where}
        GROUP BY p.forma_pago ORDER BY total DESC
    """, params)

    return {
        "titulo": "Reporte de Ventas",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "resumen": resumen,
        "detalle": detalle,
        "por_dia": por_dia,
        "por_vendedor": por_vendedor,
        "por_forma_pago": por_forma_pago,
    }


# ══════════════════════════════════════════════════════════════
#  2. PRODUCTOS FACTURADOS
# ══════════════════════════════════════════════════════════════

@router.get("/productos-facturados")
def reporte_productos_facturados(
    fecha_ini: str, fecha_fin: str,
    cliente_id:   Optional[int] = None,
    marca_id:     Optional[int] = None,
    categoria_id: Optional[int] = None,
    vendedor_id:  Optional[int] = None,
    agrupar:      str = "producto",
    u=Depends(get_current_user)
):
    conds  = ["f.estado='EMITIDA'", "f.fecha_emision BETWEEN %s AND %s"]
    params = [fecha_ini, fecha_fin]
    if cliente_id:   conds.append("f.cliente_id=%s");   params.append(cliente_id)
    if marca_id:     conds.append("p.marca_id=%s");     params.append(marca_id)
    if categoria_id: conds.append("p.categoria_id=%s"); params.append(categoria_id)
    if vendedor_id:  conds.append("f.vendedor_id=%s");  params.append(vendedor_id)
    where = "WHERE " + " AND ".join(conds)

    resumen = query_one(f"""
        SELECT COUNT(DISTINCT f.id) as facturas,
               COUNT(DISTINCT fd.producto_id) as productos_distintos,
               COALESCE(SUM(fd.cantidad),0) as unidades,
               COALESCE(SUM(fd.total),0) as total
        FROM ven_factura_detalles fd
        JOIN ven_facturas f    ON f.id=fd.factura_id
        JOIN inv_productos p   ON p.id=fd.producto_id
        {where}
    """, params)

    if agrupar == "cliente":
        detalle = query(f"""
            SELECT c.razon_social as grupo, c.identificacion as ruc,
                   COUNT(DISTINCT f.id) as facturas,
                   SUM(fd.cantidad) as unidades, SUM(fd.total) as total
            FROM ven_factura_detalles fd
            JOIN ven_facturas f  ON f.id=fd.factura_id
            JOIN inv_productos p ON p.id=fd.producto_id
            JOIN ven_clientes c  ON c.id=f.cliente_id
            {where}
            GROUP BY c.razon_social, c.identificacion
            ORDER BY total DESC
        """, params)
    elif agrupar == "marca":
        detalle = query(f"""
            SELECT COALESCE(m.nombre,'Sin marca') as grupo,
                   COUNT(DISTINCT p.id) as productos,
                   SUM(fd.cantidad) as unidades, SUM(fd.total) as total
            FROM ven_factura_detalles fd
            JOIN ven_facturas f    ON f.id=fd.factura_id
            JOIN inv_productos p   ON p.id=fd.producto_id
            LEFT JOIN inv_marcas m ON m.id=p.marca_id
            {where}
            GROUP BY m.nombre ORDER BY total DESC
        """, params)
    elif agrupar == "categoria":
        detalle = query(f"""
            SELECT COALESCE(cat.nombre,'Sin categoría') as grupo,
                   COUNT(DISTINCT p.id) as productos,
                   SUM(fd.cantidad) as unidades, SUM(fd.total) as total
            FROM ven_factura_detalles fd
            JOIN ven_facturas f       ON f.id=fd.factura_id
            JOIN inv_productos p      ON p.id=fd.producto_id
            LEFT JOIN inv_categorias cat ON cat.id=p.categoria_id
            {where}
            GROUP BY cat.nombre ORDER BY total DESC
        """, params)
    elif agrupar == "vendedor":
        detalle = query(f"""
            SELECT COALESCE(v.nombre,'Sin vendedor') as grupo,
                   COUNT(DISTINCT f.id) as facturas,
                   SUM(fd.cantidad) as unidades, SUM(fd.total) as total
            FROM ven_factura_detalles fd
            JOIN ven_facturas f        ON f.id=fd.factura_id
            JOIN inv_productos p       ON p.id=fd.producto_id
            LEFT JOIN ven_vendedores v ON v.id=f.vendedor_id
            {where}
            GROUP BY v.nombre ORDER BY total DESC
        """, params)
    else:
        detalle = query(f"""
            SELECT p.codigo, p.descripcion as grupo,
                   COALESCE(m.nombre,'') as marca,
                   COALESCE(cat.nombre,'') as categoria,
                   SUM(fd.cantidad) as unidades,
                   ROUND(AVG(fd.precio_unitario)::numeric,2) as precio_prom,
                   SUM(fd.total) as total
            FROM ven_factura_detalles fd
            JOIN ven_facturas f           ON f.id=fd.factura_id
            JOIN inv_productos p          ON p.id=fd.producto_id
            LEFT JOIN inv_marcas m        ON m.id=p.marca_id
            LEFT JOIN inv_categorias cat  ON cat.id=p.categoria_id
            {where}
            GROUP BY p.codigo, p.descripcion, m.nombre, cat.nombre
            ORDER BY total DESC
        """, params)

    return {
        "titulo": f"Productos Facturados (por {agrupar})",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "agrupar": agrupar,
        "resumen": resumen,
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  3. INVENTARIO
# ══════════════════════════════════════════════════════════════

@router.get("/inventario")
def reporte_inventario(
    bodega_id:    Optional[int] = None,
    marca_id:     Optional[int] = None,
    categoria_id: Optional[int] = None,
    solo_stock:   bool = False,
    u=Depends(get_current_user)
):
    conds  = ["p.activo=true"]
    params = []
    if marca_id:     conds.append("p.marca_id=%s");     params.append(marca_id)
    if categoria_id: conds.append("p.categoria_id=%s"); params.append(categoria_id)
    if solo_stock:   conds.append("COALESCE(st.cantidad,0) > 0")
    where = "WHERE " + " AND ".join(conds)

    bod_join = ""
    if bodega_id:
        bod_join = "AND s.bodega_id=%s"
        params_stock = [bodega_id]
    else:
        params_stock = []

    detalle = query(f"""
        SELECT p.codigo, p.descripcion,
               COALESCE(m.nombre,'') as marca,
               COALESCE(cat.nombre,'') as categoria,
               COALESCE(st.cantidad,0) as stock,
               COALESCE(co.costo,0) as costo,
               COALESCE(pr.precio,0) as precio,
               ROUND(COALESCE(st.cantidad,0) * COALESCE(co.costo,0), 2) as valor_costo,
               ROUND(COALESCE(st.cantidad,0) * COALESCE(pr.precio,0), 2) as valor_venta
        FROM inv_productos p
        LEFT JOIN inv_marcas m       ON m.id=p.marca_id
        LEFT JOIN inv_categorias cat ON cat.id=p.categoria_id
        LEFT JOIN (
            SELECT producto_id, SUM(cantidad) as cantidad
            FROM inv_stock {"WHERE bodega_id=%s" if bodega_id else ""}
            GROUP BY producto_id
        ) st ON st.producto_id=p.id
        LEFT JOIN inv_costos co ON co.producto_id=p.id
        LEFT JOIN inv_precios pr ON pr.producto_id=p.id
            AND pr.tipo_precio_id=1 AND pr.activo=true
        {where}
        ORDER BY p.descripcion
    """, (params_stock + params) if bodega_id else params)

    resumen = {
        "total_productos": len(detalle),
        "con_stock": sum(1 for d in detalle if float(d.get("stock",0)) > 0),
        "sin_stock": sum(1 for d in detalle if float(d.get("stock",0)) == 0),
        "valor_costo_total": round(sum(float(d.get("valor_costo",0)) for d in detalle), 2),
        "valor_venta_total": round(sum(float(d.get("valor_venta",0)) for d in detalle), 2),
        "unidades_total": round(sum(float(d.get("stock",0)) for d in detalle), 2),
    }

    return {
        "titulo": "Reporte de Inventario",
        "resumen": resumen,
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  4. CXC AGING (Cartera por cobrar)
# ══════════════════════════════════════════════════════════════

@router.get("/cxc-aging")
def reporte_cxc_aging(
    sucursal_id: Optional[int] = None,
    u=Depends(get_current_user)
):
    conds  = ["cxc.saldo > 0"]
    params = []
    if sucursal_id:
        conds.append("f.sucursal_id=%s"); params.append(sucursal_id)
    where = "WHERE " + " AND ".join(conds)

    detalle = query(f"""
        SELECT c.razon_social as cliente, c.identificacion as ruc,
               c.telefono, c.email,
               f.numero_factura,
               cxc.fecha_emision, cxc.fecha_vencimiento,
               CAST(cxc.valor_total AS FLOAT) as monto,
               CAST(cxc.saldo AS FLOAT) as saldo,
               (CURRENT_DATE - cxc.fecha_vencimiento) as dias_vencido,
               CASE
                   WHEN cxc.fecha_vencimiento >= CURRENT_DATE THEN 'VIGENTE'
                   WHEN (CURRENT_DATE - cxc.fecha_vencimiento) <= 30 THEN '1-30'
                   WHEN (CURRENT_DATE - cxc.fecha_vencimiento) <= 60 THEN '31-60'
                   WHEN (CURRENT_DATE - cxc.fecha_vencimiento) <= 90 THEN '61-90'
                   ELSE '90+'
               END as rango
        FROM fin_cxc cxc
        JOIN ven_clientes c     ON c.id=cxc.cliente_id
        LEFT JOIN ven_facturas f ON f.id=cxc.factura_id
        {where}
        ORDER BY cxc.fecha_vencimiento
    """, params)

    vigente = sum(float(d["saldo"]) for d in detalle if d["rango"]=="VIGENTE")
    r30  = sum(float(d["saldo"]) for d in detalle if d["rango"]=="1-30")
    r60  = sum(float(d["saldo"]) for d in detalle if d["rango"]=="31-60")
    r90  = sum(float(d["saldo"]) for d in detalle if d["rango"]=="61-90")
    r90p = sum(float(d["saldo"]) for d in detalle if d["rango"]=="90+")

    return {
        "titulo": "Cartera por Cobrar — Aging",
        "resumen": {
            "total_cartera": round(vigente+r30+r60+r90+r90p, 2),
            "vigente": round(vigente, 2),
            "1_30": round(r30, 2),
            "31_60": round(r60, 2),
            "61_90": round(r90, 2),
            "90_mas": round(r90p, 2),
            "cuentas": len(detalle),
            "clientes": len(set(d["cliente"] for d in detalle)),
        },
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  5. COMPRAS
# ══════════════════════════════════════════════════════════════

@router.get("/compras")
def reporte_compras(
    fecha_ini: str, fecha_fin: str,
    proveedor_id: Optional[int] = None,
    u=Depends(get_current_user)
):
    conds  = ["c.estado='CONFIRMADA'", "DATE(c.fecha) BETWEEN %s AND %s"]
    params = [fecha_ini, fecha_fin]
    if proveedor_id: conds.append("c.proveedor_id=%s"); params.append(proveedor_id)
    where = "WHERE " + " AND ".join(conds)

    resumen = query_one(f"""
        SELECT COUNT(*) as compras,
               COALESCE(SUM(c.subtotal_0),0) as subtotal_0,
               COALESCE(SUM(c.subtotal_iva),0) as subtotal_iva,
               COALESCE(SUM(c.iva),0) as iva,
               COALESCE(SUM(c.total),0) as total
        FROM com_compras c {where}
    """, params)

    detalle = query(f"""
        SELECT c.id, c.num_documento, DATE(c.fecha) as fecha,
               p.razon_social as proveedor, p.identificacion as ruc,
               c.subtotal_0, c.subtotal_iva, c.iva, c.total
        FROM com_compras c
        JOIN com_proveedores p ON p.id=c.proveedor_id
        {where}
        ORDER BY c.fecha, c.id
    """, params)

    por_proveedor = query(f"""
        SELECT p.razon_social as proveedor,
               COUNT(*) as compras, SUM(c.total) as total
        FROM com_compras c
        JOIN com_proveedores p ON p.id=c.proveedor_id
        {where}
        GROUP BY p.razon_social ORDER BY total DESC
    """, params)

    return {
        "titulo": "Reporte de Compras",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "resumen": resumen,
        "detalle": detalle,
        "por_proveedor": por_proveedor,
    }


# ══════════════════════════════════════════════════════════════
#  6. COMISIONES VENDEDORES
# ══════════════════════════════════════════════════════════════

@router.get("/comisiones")
def reporte_comisiones(
    mes: Optional[str] = None,
    u=Depends(get_current_user)
):
    if not mes:
        from datetime import date
        mes = date.today().strftime("%Y-%m")

    detalle = query("""
        SELECT v.id, v.codigo, v.nombre, v.apellidos,
               CAST(COALESCE(v.comision_pct,0) AS FLOAT) as comision_pct,
               CAST(COALESCE(v.meta_mensual,0) AS FLOAT) as meta,
               s.nombre as sucursal,
               CAST(COALESCE(ventas.base_imp,0) AS FLOAT) as base_imponible,
               CAST(COALESCE(ventas.total,0) AS FLOAT) as total_ventas,
               COALESCE(ventas.facturas,0) as facturas,
               ROUND(CAST(COALESCE(ventas.base_imp,0) * COALESCE(v.comision_pct,0) / 100 AS NUMERIC), 2) as comision,
               CASE WHEN v.meta_mensual > 0
                   THEN ROUND(CAST(COALESCE(ventas.base_imp,0) / v.meta_mensual * 100 AS NUMERIC), 1)
                   ELSE 0
               END as pct_cumplimiento
        FROM ven_vendedores v
        LEFT JOIN sys_sucursales s ON s.id=v.sucursal_id
        LEFT JOIN (
            SELECT f.vendedor_id,
                   SUM(f.subtotal_0 + f.subtotal_iva) as base_imp,
                   SUM(f.total) as total,
                   COUNT(*) as facturas
            FROM ven_facturas f
            WHERE f.estado='EMITIDA'
              AND DATE_TRUNC('month', f.fecha_emision) = DATE_TRUNC('month', %s::date)
            GROUP BY f.vendedor_id
        ) ventas ON ventas.vendedor_id=v.id
        WHERE v.activo=true
        ORDER BY COALESCE(ventas.base_imp,0) DESC
    """, (mes + "-01",))

    total_ventas = sum(float(d.get("total_ventas",0)) for d in detalle)
    total_comisiones = sum(float(d.get("comision",0)) for d in detalle)

    return {
        "titulo": f"Comisiones Vendedores — {mes}",
        "mes": mes,
        "resumen": {
            "vendedores": len(detalle),
            "total_ventas": round(total_ventas, 2),
            "total_comisiones": round(total_comisiones, 2),
        },
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  7. CAJA (Resumen de sesiones)
# ══════════════════════════════════════════════════════════════

@router.get("/caja")
def reporte_caja(
    fecha_ini: str, fecha_fin: str,
    caja_id: Optional[int] = None,
    u=Depends(get_current_user)
):
    conds  = ["DATE(cs.fecha_apertura) BETWEEN %s AND %s"]
    params = [fecha_ini, fecha_fin]
    if caja_id: conds.append("cs.caja_id=%s"); params.append(caja_id)
    where = "WHERE " + " AND ".join(conds)

    detalle = query(f"""
        SELECT cs.id, c.nombre as caja,
               DATE(cs.fecha_apertura) as fecha,
               u.nombre as usuario,
               CAST(COALESCE(cs.monto_apertura,0) AS FLOAT) as apertura,
               CAST(COALESCE(cs.total_efectivo_sistema,0) AS FLOAT) as efectivo,
               CAST(COALESCE(cs.total_tarjeta_sistema,0) AS FLOAT) as tarjeta,
               CAST(COALESCE(cs.total_transferencia_sistema,0) AS FLOAT) as transferencia,
               CAST(COALESCE(cs.total_contado,0) AS FLOAT) as contado,
               CAST(COALESCE(cs.diferencia,0) AS FLOAT) as diferencia,
               cs.estado
        FROM caj_sesiones cs
        JOIN caj_cajas c       ON c.id=cs.caja_id
        LEFT JOIN sys_usuarios u ON u.id=cs.usuario_id
        {where}
        ORDER BY cs.fecha_apertura DESC
    """, params)

    return {
        "titulo": "Reporte de Caja",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "resumen": {
            "sesiones": len(detalle),
            "total_efectivo": round(sum(float(d.get("efectivo",0)) for d in detalle), 2),
            "total_tarjeta": round(sum(float(d.get("tarjeta",0)) for d in detalle), 2),
            "total_diferencia": round(sum(float(d.get("diferencia",0)) for d in detalle), 2),
        },
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  EXPORTACIÓN — Excel y PDF genéricos
# ══════════════════════════════════════════════════════════════

@router.post("/exportar/excel")
async def exportar_excel(request: Request, u=Depends(get_current_user)):
    body = await request.json()
    """
    Exporta cualquier reporte a Excel.
    Body: { titulo, columnas: [{key, label, width}], filas: [{...}], resumen: {...} }
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (body.get("titulo") or "Reporte")[:31]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a365d")
    total_font  = Font(bold=True, size=11)
    total_fill  = PatternFill("solid", fgColor="e2e8f0")
    border      = Border(
        bottom=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
    )

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=max(len(body.get("columnas", [])), 1))
    ws.cell(1, 1, body.get("titulo", "Reporte")).font = Font(bold=True, size=14)
    if body.get("fecha_ini"):
        ws.cell(2, 1, f"Período: {body['fecha_ini']} al {body.get('fecha_fin','')}")

    row = 4

    # Resumen
    if body.get("resumen"):
        for k, v in body["resumen"].items():
            ws.cell(row, 1, k.replace("_", " ").title()).font = Font(bold=True)
            ws.cell(row, 2, v)
            row += 1
        row += 1

    # Columnas
    columnas = body.get("columnas", [])
    for i, col in enumerate(columnas, 1):
        cell = ws.cell(row, i, col.get("label", col.get("key", "")))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        if col.get("width"):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col["width"]
    row += 1

    # Filas
    for fila in body.get("filas", []):
        for i, col in enumerate(columnas, 1):
            val = fila.get(col.get("key", ""), "")
            cell = ws.cell(row, i, val)
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00' if isinstance(val, float) else '#,##0'
        row += 1

    # Totales
    if body.get("totales"):
        for i, col in enumerate(columnas, 1):
            key = col.get("key", "")
            val = body["totales"].get(key, "")
            cell = ws.cell(row, i, val)
            cell.font = total_font
            cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    nombre = (body.get("titulo") or "reporte").replace(" ", "_")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}.xlsx"'}
    )


@router.post("/exportar/pdf")
async def exportar_pdf(request: Request, u=Depends(get_current_user)):
    body = await request.json()
    """
    Exporta cualquier reporte a PDF.
    Body: { titulo, columnas: [{key, label}], filas: [{...}], resumen: {...} }
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    columnas = body.get("columnas", [])
    usar_landscape = len(columnas) > 6

    doc = SimpleDocTemplate(buf,
        pagesize=landscape(A4) if usar_landscape else A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=10*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph(f"<b>{body.get('titulo','Reporte')}</b>", styles["Title"]))
    if body.get("fecha_ini"):
        elements.append(Paragraph(
            f"Período: {body['fecha_ini']} al {body.get('fecha_fin','')}",
            styles["Normal"]))
    elements.append(Spacer(1, 5*mm))

    # Resumen
    if body.get("resumen"):
        res_data = [[k.replace("_"," ").title(), str(v)]
                     for k, v in body["resumen"].items()]
        res_table = Table(res_data, colWidths=[120, 120])
        res_table.setStyle(TableStyle([
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ]))
        elements.append(res_table)
        elements.append(Spacer(1, 5*mm))

    # Tabla de datos
    if columnas and body.get("filas"):
        header = [col.get("label", col.get("key","")) for col in columnas]
        data = [header]
        for fila in body["filas"]:
            row = []
            for col in columnas:
                val = fila.get(col.get("key",""), "")
                if isinstance(val, float):
                    val = f"{val:,.2f}"
                elif val is None:
                    val = ""
                row.append(str(val)[:40])
            data.append(row)

        # Totales
        if body.get("totales"):
            tot_row = []
            for col in columnas:
                val = body["totales"].get(col.get("key",""), "")
                if isinstance(val, float):
                    val = f"{val:,.2f}"
                tot_row.append(str(val))
            data.append(tot_row)

        page_w = (landscape(A4) if usar_landscape else A4)[0] - 30*mm
        col_w = page_w / len(columnas)

        t = Table(data, colWidths=[col_w]*len(columnas), repeatRows=1)
        style = [
            ("FONTSIZE",    (0,0), (-1,-1), 7),
            ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
            ("BACKGROUND",  (0,0), (-1,0),  colors.HexColor("#1a365d")),
            ("TEXTCOLOR",   (0,0), (-1,0),  colors.white),
            ("ALIGN",       (0,0), (-1,0),  "CENTER"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f7fafc")]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#cbd5e0")),
            ("TOPPADDING",  (0,0), (-1,-1), 2),
            ("BOTTOMPADDING",(0,0),(-1,-1), 2),
        ]
        if body.get("totales"):
            style.append(("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"))
            style.append(("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#e2e8f0")))
        t.setStyle(TableStyle(style))
        elements.append(t)

    doc.build(elements)
    nombre = (body.get("titulo") or "reporte").replace(" ", "_")
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nombre}.pdf"'}
    )


# ══════════════════════════════════════════════════════════════
#  8. ATS — Anexo Transaccional Simplificado
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
#  FORMULARIO 104 — IVA
# ══════════════════════════════════════════════════════════════

@router.get("/formulario-104")
def formulario_104(mes: str, u=Depends(get_current_user)):
    """Pre-fills form 104 (IVA) data from transactions."""
    year, month = mes.split('-')
    import calendar
    last_day = calendar.monthrange(int(year), int(month))[1]
    fi, ff = f"{mes}-01", f"{mes}-{last_day}"

    # Ventas gravadas con tarifa
    ventas_iva = query_one("""
        SELECT COALESCE(SUM(subtotal_iva),0) as base, COALESCE(SUM(iva),0) as impuesto
        FROM ven_facturas WHERE estado='EMITIDA' AND fecha_emision BETWEEN %s AND %s
    """, (fi, ff))

    # Ventas 0%
    ventas_0 = query_one("""
        SELECT COALESCE(SUM(subtotal_0),0) as base
        FROM ven_facturas WHERE estado='EMITIDA' AND fecha_emision BETWEEN %s AND %s
    """, (fi, ff))

    # Compras gravadas
    compras_iva = query_one("""
        SELECT COALESCE(SUM(subtotal_iva),0) as base, COALESCE(SUM(iva),0) as impuesto
        FROM com_compras WHERE estado='CONFIRMADA' AND DATE(fecha) BETWEEN %s AND %s
    """, (fi, ff))

    # Retenciones de IVA que nos hicieron
    ret_iva_recibidas = query_one("""
        SELECT COALESCE(SUM(d.valor_retenido),0) as total
        FROM sri_retencion_detalles d
        JOIN sri_retenciones r ON r.id=d.retencion_id
        WHERE r.tipo='RECIBIDA' AND d.tipo_impuesto='IVA'
          AND r.fecha_emision BETWEEN %s AND %s
    """, (fi, ff))

    iva_causado = float(ventas_iva['impuesto'])
    credito_tributario = float(compras_iva['impuesto'])
    ret_recibidas = float(ret_iva_recibidas['total'])
    impuesto_a_pagar = max(0, iva_causado - credito_tributario - ret_recibidas)

    return {
        "periodo": mes, "formulario": "104",
        "ventas_tarifa_iva": {"base": float(ventas_iva['base']), "impuesto": float(ventas_iva['impuesto'])},
        "ventas_tarifa_0": {"base": float(ventas_0['base'])},
        "compras_tarifa_iva": {"base": float(compras_iva['base']), "impuesto": float(compras_iva['impuesto'])},
        "credito_tributario": credito_tributario,
        "retenciones_iva_recibidas": ret_recibidas,
        "iva_causado": iva_causado,
        "impuesto_a_pagar": round(impuesto_a_pagar, 2),
    }


# ══════════════════════════════════════════════════════════════
#  FORMULARIO 103 — Retenciones en la Fuente
# ══════════════════════════════════════════════════════════════

@router.get("/formulario-103")
def formulario_103(mes: str, u=Depends(get_current_user)):
    """Pre-fills form 103 (Retenciones en la Fuente) data."""
    year, month = mes.split('-')
    import calendar
    last_day = calendar.monthrange(int(year), int(month))[1]
    fi, ff = f"{mes}-01", f"{mes}-{last_day}"

    retenciones = query("""
        SELECT d.codigo_retencion, d.porcentaje,
               SUM(d.base_imponible) as base_imponible,
               SUM(d.valor_retenido) as valor_retenido,
               COUNT(*) as num_comprobantes
        FROM sri_retencion_detalles d
        JOIN sri_retenciones r ON r.id=d.retencion_id
        WHERE r.tipo='EMITIDA' AND d.tipo_impuesto='RENTA'
          AND r.fecha_emision BETWEEN %s AND %s AND r.estado='EMITIDA'
        GROUP BY d.codigo_retencion, d.porcentaje
        ORDER BY d.codigo_retencion
    """, (fi, ff))

    total = sum(float(r['valor_retenido']) for r in retenciones)

    return {
        "periodo": mes, "formulario": "103",
        "retenciones": retenciones,
        "total_retenido": round(total, 2),
        "num_codigos": len(retenciones),
    }


# ══════════════════════════════════════════════════════════════
#  8. ATS — Anexo Transaccional Simplificado
# ══════════════════════════════════════════════════════════════

@router.get("/ats")
def generar_ats(mes: str, u=Depends(get_current_user)):
    """
    Generates ATS data for a month (format: YYYY-MM).
    Returns: compras (purchases with retentions), ventas (sales),
    anulados (voided documents), totals.
    """
    import calendar as cal_mod
    year, month = mes.split('-')
    fecha_ini = f"{mes}-01"
    last_day = cal_mod.monthrange(int(year), int(month))[1]
    fecha_fin = f"{mes}-{last_day}"

    # Ventas del periodo
    ventas = query("""
        SELECT f.numero_factura, f.fecha_emision, f.clave_acceso,
               c.tipo_identificacion, c.identificacion, c.razon_social,
               f.subtotal_0, f.subtotal_iva, f.iva, f.total,
               f.estado
        FROM ven_facturas f
        JOIN ven_clientes c ON c.id=f.cliente_id
        WHERE f.estado='EMITIDA' AND f.fecha_emision BETWEEN %s AND %s
        ORDER BY f.fecha_emision
    """, (fecha_ini, fecha_fin))

    # Compras del periodo
    compras = query("""
        SELECT c.num_documento, c.fecha, c.subtotal_0, c.subtotal_iva,
               c.iva, c.total, p.tipo_identificacion, p.identificacion,
               p.razon_social, c.estado
        FROM com_compras c
        JOIN com_proveedores p ON p.id=c.proveedor_id
        WHERE c.estado='CONFIRMADA' AND DATE(c.fecha) BETWEEN %s AND %s
        ORDER BY c.fecha
    """, (fecha_ini, fecha_fin))

    # Retenciones emitidas
    retenciones = query("""
        SELECT r.numero, r.fecha_emision, r.total_retenido, r.clave_acceso,
               p.identificacion, p.razon_social
        FROM sri_retenciones r
        LEFT JOIN com_proveedores p ON p.id=r.proveedor_id
        WHERE r.tipo='EMITIDA' AND r.estado='EMITIDA'
          AND r.fecha_emision BETWEEN %s AND %s
        ORDER BY r.fecha_emision
    """, (fecha_ini, fecha_fin))

    # Notas de credito
    ncs = query("""
        SELECT d.numero, d.fecha, d.total, c.identificacion, c.razon_social
        FROM ven_devoluciones d
        JOIN ven_clientes c ON c.id=d.cliente_id
        WHERE d.estado='EMITIDA' AND DATE(d.fecha) BETWEEN %s AND %s
    """, (fecha_ini, fecha_fin))

    # Documentos anulados
    anulados = query("""
        SELECT numero_factura as numero, '01' as tipo_doc
        FROM ven_facturas WHERE estado='ANULADA' AND fecha_emision BETWEEN %s AND %s
        UNION ALL
        SELECT numero as numero, '04' as tipo_doc
        FROM ven_devoluciones WHERE estado='ANULADA' AND DATE(fecha) BETWEEN %s AND %s
    """, (fecha_ini, fecha_fin, fecha_ini, fecha_fin))

    return {
        "periodo": mes,
        "ventas": ventas,
        "compras": compras,
        "retenciones": retenciones,
        "notas_credito": ncs,
        "anulados": anulados,
        "resumen": {
            "total_ventas": sum(float(v.get("total", 0)) for v in ventas),
            "total_compras": sum(float(c.get("total", 0)) for c in compras),
            "total_retenciones": sum(float(r.get("total_retenido", 0)) for r in retenciones),
            "num_ventas": len(ventas),
            "num_compras": len(compras),
            "num_retenciones": len(retenciones),
            "num_anulados": len(anulados),
        }
    }


# ══════════════════════════════════════════════════════════════
#  9. RENTABILIDAD POR PRODUCTO
# ══════════════════════════════════════════════════════════════

@router.get("/rentabilidad")
def reporte_rentabilidad(fecha_ini: str, fecha_fin: str, u=Depends(get_current_user)):
    """Product profitability: revenue - cost = margin."""
    detalle = query("""
        SELECT p.codigo, p.descripcion,
               COALESCE(m.nombre,'') as marca, COALESCE(cat.nombre,'') as categoria,
               SUM(fd.cantidad) as unidades,
               SUM(fd.total) as ingreso,
               SUM(fd.cantidad * COALESCE(co.costo,0)) as costo_total,
               SUM(fd.total) - SUM(fd.cantidad * COALESCE(co.costo,0)) as utilidad,
               CASE WHEN SUM(fd.total) > 0
                   THEN ROUND(((SUM(fd.total) - SUM(fd.cantidad * COALESCE(co.costo,0))) / SUM(fd.total) * 100)::numeric, 1)
                   ELSE 0 END as margen_pct
        FROM ven_factura_detalles fd
        JOIN inv_productos p ON p.id=fd.producto_id
        JOIN ven_facturas f ON f.id=fd.factura_id
        LEFT JOIN inv_costos co ON co.producto_id=p.id
        LEFT JOIN inv_marcas m ON m.id=p.marca_id
        LEFT JOIN inv_categorias cat ON cat.id=p.categoria_id
        WHERE f.estado='EMITIDA' AND f.fecha_emision BETWEEN %s AND %s
        GROUP BY p.codigo, p.descripcion, m.nombre, cat.nombre
        ORDER BY utilidad DESC
    """, (fecha_ini, fecha_fin))

    total_ingreso = sum(float(d.get('ingreso',0)) for d in detalle)
    total_costo = sum(float(d.get('costo_total',0)) for d in detalle)
    total_utilidad = sum(float(d.get('utilidad',0)) for d in detalle)
    margen_global = round((total_utilidad / total_ingreso * 100), 1) if total_ingreso > 0 else 0

    return {
        "titulo": "Rentabilidad por Producto",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "resumen": {"total_ingreso": round(total_ingreso,2), "total_costo": round(total_costo,2),
                     "total_utilidad": round(total_utilidad,2), "margen_global": margen_global,
                     "productos": len(detalle)},
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  10. CLIENTES MAS RENTABLES (PARETO 80/20)
# ══════════════════════════════════════════════════════════════

@router.get("/clientes-rentables")
def reporte_clientes_rentables(fecha_ini: str, fecha_fin: str, u=Depends(get_current_user)):
    """Top customers by revenue with Pareto (80/20) analysis."""
    detalle = query("""
        SELECT c.razon_social as cliente, c.identificacion as ruc,
               COUNT(DISTINCT f.id) as facturas,
               SUM(f.total) as total,
               SUM(f.subtotal_iva + f.subtotal_0) as base_imponible
        FROM ven_facturas f
        JOIN ven_clientes c ON c.id=f.cliente_id
        WHERE f.estado='EMITIDA' AND f.fecha_emision BETWEEN %s AND %s
        GROUP BY c.razon_social, c.identificacion
        ORDER BY total DESC
    """, (fecha_ini, fecha_fin))

    gran_total = sum(float(d['total']) for d in detalle)
    acumulado = 0
    for d in detalle:
        acumulado += float(d['total'])
        d['pct_total'] = round(float(d['total']) / gran_total * 100, 1) if gran_total > 0 else 0
        d['pct_acumulado'] = round(acumulado / gran_total * 100, 1) if gran_total > 0 else 0
        d['es_pareto'] = d['pct_acumulado'] <= 80

    clientes_pareto = sum(1 for d in detalle if d['es_pareto'])

    return {
        "titulo": "Clientes Más Rentables (Pareto 80/20)",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "resumen": {"total_clientes": len(detalle), "total_ventas": round(gran_total,2),
                     "clientes_pareto": clientes_pareto,
                     "pct_clientes_pareto": round(clientes_pareto/len(detalle)*100,1) if detalle else 0},
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  11. STOCK MUERTO (sin movimiento)
# ══════════════════════════════════════════════════════════════

@router.get("/stock-muerto")
def reporte_stock_muerto(dias: int = 90, u=Depends(get_current_user)):
    """Products with stock but no sales in the last N days."""
    detalle = query("""
        SELECT p.codigo, p.descripcion,
               COALESCE(m.nombre,'') as marca, COALESCE(cat.nombre,'') as categoria,
               COALESCE(SUM(s.cantidad),0) as stock,
               COALESCE(co.costo,0) as costo_unitario,
               COALESCE(SUM(s.cantidad),0) * COALESCE(co.costo,0) as valor_costo,
               (SELECT MAX(f.fecha_emision) FROM ven_factura_detalles fd
                JOIN ven_facturas f ON f.id=fd.factura_id
                WHERE fd.producto_id=p.id AND f.estado='EMITIDA') as ultima_venta,
               CURRENT_DATE - COALESCE(
                   (SELECT MAX(f.fecha_emision) FROM ven_factura_detalles fd
                    JOIN ven_facturas f ON f.id=fd.factura_id
                    WHERE fd.producto_id=p.id AND f.estado='EMITIDA'),
                   p.created_at::date
               ) as dias_sin_venta
        FROM inv_productos p
        LEFT JOIN inv_stock s ON s.producto_id=p.id
        LEFT JOIN inv_costos co ON co.producto_id=p.id
        LEFT JOIN inv_marcas m ON m.id=p.marca_id
        LEFT JOIN inv_categorias cat ON cat.id=p.categoria_id
        WHERE p.activo=true
        GROUP BY p.id, p.codigo, p.descripcion, m.nombre, cat.nombre, co.costo, p.created_at
        HAVING COALESCE(SUM(s.cantidad),0) > 0
           AND (CURRENT_DATE - COALESCE(
               (SELECT MAX(f.fecha_emision)::date FROM ven_factura_detalles fd
                JOIN ven_facturas f ON f.id=fd.factura_id
                WHERE fd.producto_id=p.id AND f.estado='EMITIDA'),
               p.created_at::date
           ))::integer >= %s
        ORDER BY dias_sin_venta DESC
    """, (dias,))

    total_valor = sum(float(d.get('valor_costo',0)) for d in detalle)
    return {
        "titulo": f"Stock Sin Movimiento (>{dias} días)",
        "dias": dias,
        "resumen": {"productos": len(detalle), "valor_total_retenido": round(total_valor,2)},
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  12. CXP AGING (Cuentas por pagar)
# ══════════════════════════════════════════════════════════════

@router.get("/cxp-aging")
def reporte_cxp_aging(u=Depends(get_current_user)):
    """Accounts payable aging report."""
    detalle = query("""
        SELECT p.razon_social as proveedor, p.identificacion as ruc, p.telefono,
               c.num_documento, cx.fecha_emision, cx.fecha_vencimiento,
               CAST(cx.valor_total AS FLOAT) as monto,
               CAST(cx.saldo AS FLOAT) as saldo,
               (CURRENT_DATE - cx.fecha_vencimiento) as dias_vencido,
               CASE
                   WHEN cx.fecha_vencimiento >= CURRENT_DATE THEN 'VIGENTE'
                   WHEN (CURRENT_DATE - cx.fecha_vencimiento) <= 30 THEN '1-30'
                   WHEN (CURRENT_DATE - cx.fecha_vencimiento) <= 60 THEN '31-60'
                   WHEN (CURRENT_DATE - cx.fecha_vencimiento) <= 90 THEN '61-90'
                   ELSE '90+'
               END as rango
        FROM fin_cxp cx
        JOIN com_proveedores p ON p.id=cx.proveedor_id
        LEFT JOIN com_compras c ON c.id=cx.compra_id
        WHERE cx.saldo > 0
        ORDER BY cx.fecha_vencimiento
    """)

    vigente = sum(float(d['saldo']) for d in detalle if d['rango']=='VIGENTE')
    r30 = sum(float(d['saldo']) for d in detalle if d['rango']=='1-30')
    r60 = sum(float(d['saldo']) for d in detalle if d['rango']=='31-60')
    r90 = sum(float(d['saldo']) for d in detalle if d['rango']=='61-90')
    r90p = sum(float(d['saldo']) for d in detalle if d['rango']=='90+')

    return {
        "titulo": "Cuentas por Pagar — Aging",
        "resumen": {"total_cartera": round(vigente+r30+r60+r90+r90p,2),
                     "vigente": round(vigente,2), "1_30": round(r30,2), "31_60": round(r60,2),
                     "61_90": round(r90,2), "90_mas": round(r90p,2),
                     "cuentas": len(detalle), "proveedores": len(set(d['proveedor'] for d in detalle))},
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  13. COMPARATIVO DE VENTAS
# ══════════════════════════════════════════════════════════════

@router.get("/comparativo-ventas")
def reporte_comparativo(periodo1_ini: str, periodo1_fin: str, periodo2_ini: str, periodo2_fin: str, u=Depends(get_current_user)):
    """Compare sales between two periods."""
    def get_ventas(fi, ff):
        return query_one("""
            SELECT COUNT(*) as facturas, COALESCE(SUM(total),0) as total,
                   COALESCE(SUM(subtotal_iva),0) as subtotal_iva,
                   COALESCE(SUM(iva),0) as iva
            FROM ven_facturas WHERE estado='EMITIDA' AND fecha_emision BETWEEN %s AND %s
        """, (fi, ff))

    p1 = get_ventas(periodo1_ini, periodo1_fin)
    p2 = get_ventas(periodo2_ini, periodo2_fin)

    total_p1 = float(p1['total'])
    total_p2 = float(p2['total'])
    variacion = round(((total_p2 - total_p1) / total_p1 * 100), 1) if total_p1 > 0 else 0

    # Daily breakdown for both periods
    dias_p1 = query("SELECT fecha_emision::date as fecha, COUNT(*) as facturas, SUM(total) as total FROM ven_facturas WHERE estado='EMITIDA' AND fecha_emision BETWEEN %s AND %s GROUP BY fecha ORDER BY fecha", (periodo1_ini, periodo1_fin))
    dias_p2 = query("SELECT fecha_emision::date as fecha, COUNT(*) as facturas, SUM(total) as total FROM ven_facturas WHERE estado='EMITIDA' AND fecha_emision BETWEEN %s AND %s GROUP BY fecha ORDER BY fecha", (periodo2_ini, periodo2_fin))

    return {
        "titulo": "Comparativo de Ventas",
        "periodo1": {"ini": periodo1_ini, "fin": periodo1_fin, **p1, "total": total_p1},
        "periodo2": {"ini": periodo2_ini, "fin": periodo2_fin, **p2, "total": total_p2},
        "variacion_pct": variacion,
        "diferencia": round(total_p2 - total_p1, 2),
        "dias_p1": dias_p1, "dias_p2": dias_p2,
    }


# ══════════════════════════════════════════════════════════════
#  14. SERVICIO TECNICO
# ══════════════════════════════════════════════════════════════

@router.get("/servicio-tecnico")
def reporte_servicio_tecnico(fecha_ini: str, fecha_fin: str, u=Depends(get_current_user)):
    """Service orders report."""
    detalle = query("""
        SELECT o.numero, DATE(o.fecha_ingreso) as fecha, c.razon_social as cliente,
               o.equipo_tipo, o.equipo_marca, o.equipo_modelo,
               t.nombre as tecnico, o.estado, o.prioridad,
               o.costo_estimado, o.costo_final, o.anticipo,
               EXTRACT(DAY FROM COALESCE(o.fecha_cierre, NOW()) - o.fecha_ingreso) as dias
        FROM srv_ordenes o
        JOIN ven_clientes c ON c.id=o.cliente_id
        LEFT JOIN srv_tecnicos t ON t.id=o.tecnico_id
        WHERE o.fecha_ingreso::date BETWEEN %s AND %s
        ORDER BY o.fecha_ingreso DESC
    """, (fecha_ini, fecha_fin))

    total_ordenes = len(detalle)
    entregadas = sum(1 for d in detalle if d['estado']=='ENTREGADO')
    en_proceso = sum(1 for d in detalle if d['estado'] not in ('ENTREGADO','CANCELADO'))
    ingresos = sum(float(d.get('costo_final',0)) for d in detalle if d['estado']=='ENTREGADO')
    promedio_dias = round(sum(float(d.get('dias',0)) for d in detalle if d['estado']=='ENTREGADO') / max(entregadas,1), 1)

    return {
        "titulo": "Reporte de Servicio Técnico",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "resumen": {"total_ordenes": total_ordenes, "entregadas": entregadas, "en_proceso": en_proceso,
                     "ingresos": round(ingresos,2), "promedio_dias": promedio_dias},
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  15. NOMINA (costos laborales)
# ══════════════════════════════════════════════════════════════

@router.get("/nomina")
def reporte_nomina(periodo: str, u=Depends(get_current_user)):
    """Payroll cost report."""
    detalle = query("""
        SELECT e.nombres || ' ' || e.apellidos as empleado, e.cedula, e.cargo, e.departamento,
               r.salario_base, r.valor_horas_extras_50, r.valor_horas_extras_100,
               r.comisiones, r.bonificaciones, r.total_ingresos,
               r.aporte_iess_personal, r.aporte_iess_patronal,
               r.prestamos_empresa, r.anticipo, r.total_descuentos,
               r.neto_a_pagar, r.decimo_tercero, r.decimo_cuarto,
               r.fondos_reserva, r.vacaciones_provision
        FROM nom_roles_pago r
        JOIN nom_empleados e ON e.id=r.empleado_id
        WHERE r.periodo=%s AND r.estado='APROBADO'
        ORDER BY e.apellidos, e.nombres
    """, (periodo,))

    total_ingresos = sum(float(d.get('total_ingresos',0)) for d in detalle)
    total_iess_patronal = sum(float(d.get('aporte_iess_patronal',0)) for d in detalle)
    total_neto = sum(float(d.get('neto_a_pagar',0)) for d in detalle)
    total_provisiones = sum(float(d.get('decimo_tercero',0))+float(d.get('decimo_cuarto',0))+float(d.get('fondos_reserva',0))+float(d.get('vacaciones_provision',0)) for d in detalle)
    costo_total = total_ingresos + total_iess_patronal + total_provisiones

    return {
        "titulo": f"Reporte de Nómina — {periodo}",
        "periodo": periodo,
        "resumen": {"empleados": len(detalle), "total_ingresos": round(total_ingresos,2),
                     "total_iess_patronal": round(total_iess_patronal,2), "total_neto": round(total_neto,2),
                     "total_provisiones": round(total_provisiones,2), "costo_total_empresa": round(costo_total,2)},
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  16. DEVOLUCIONES / NOTAS DE CREDITO
# ══════════════════════════════════════════════════════════════

@router.get("/devoluciones")
def reporte_devoluciones(fecha_ini: str, fecha_fin: str, u=Depends(get_current_user)):
    detalle = query("""
        SELECT d.numero, DATE(d.fecha) as fecha, c.razon_social as cliente,
               c.identificacion as ruc, f.numero_factura as factura_ref,
               d.motivo, d.total, d.estado
        FROM ven_devoluciones d
        JOIN ven_clientes c ON c.id=d.cliente_id
        LEFT JOIN ven_facturas f ON f.id=d.factura_id
        WHERE d.estado='EMITIDA' AND DATE(d.fecha) BETWEEN %s AND %s
        ORDER BY d.fecha DESC
    """, (fecha_ini, fecha_fin))
    total = sum(float(d['total']) for d in detalle)
    return {
        "titulo": "Reporte de Devoluciones / Notas de Crédito",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "resumen": {"total_devoluciones": len(detalle), "monto_total": round(total,2)},
        "detalle": detalle,
    }


# ══════════════════════════════════════════════════════════════
#  17. DASHBOARD EJECUTIVO
# ══════════════════════════════════════════════════════════════

@router.get("/dashboard-ejecutivo")
def dashboard_ejecutivo(fecha_ini: str, fecha_fin: str, u=Depends(get_current_user)):
    """Executive summary with all key metrics."""
    ventas = query_one("SELECT COUNT(*) as n, COALESCE(SUM(total),0) as total FROM ven_facturas WHERE estado='EMITIDA' AND fecha_emision BETWEEN %s AND %s", (fecha_ini, fecha_fin))
    compras = query_one("SELECT COUNT(*) as n, COALESCE(SUM(total),0) as total FROM com_compras WHERE estado='CONFIRMADA' AND DATE(fecha) BETWEEN %s AND %s", (fecha_ini, fecha_fin))
    cxc = query_one("SELECT COALESCE(SUM(saldo),0) as total FROM fin_cxc WHERE saldo>0")
    cxp = query_one("SELECT COALESCE(SUM(saldo),0) as total FROM fin_cxp WHERE saldo>0")
    inventario = query_one("SELECT COALESCE(SUM(s.cantidad*COALESCE(c.costo,0)),0) as total FROM inv_stock s LEFT JOIN inv_costos c ON c.producto_id=s.producto_id")

    try:
        nomina = query_one("SELECT COALESCE(SUM(total_ingresos+aporte_iess_patronal),0) as costo FROM nom_roles_pago WHERE estado='APROBADO' AND periodo BETWEEN %s AND %s", (fecha_ini[:7], fecha_fin[:7]))
        costo_nomina = float(nomina['costo']) if nomina else 0
    except: costo_nomina = 0

    utilidad_bruta = float(ventas['total']) - float(compras['total'])
    utilidad_neta = utilidad_bruta - costo_nomina

    return {
        "titulo": "Dashboard Ejecutivo",
        "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
        "ventas": {"facturas": ventas['n'], "total": float(ventas['total'])},
        "compras": {"compras": compras['n'], "total": float(compras['total'])},
        "utilidad_bruta": round(utilidad_bruta, 2),
        "costo_nomina": round(costo_nomina, 2),
        "utilidad_neta": round(utilidad_neta, 2),
        "cxc_pendiente": float(cxc['total']),
        "cxp_pendiente": float(cxp['total']),
        "valor_inventario": float(inventario['total']),
        "margen_bruto_pct": round(utilidad_bruta / float(ventas['total']) * 100, 1) if float(ventas['total']) > 0 else 0,
        "resumen": {},
        "detalle": [],
    }


# ══════════════════════════════════════════════════════════════
#  18. FLUJO DE CAJA PROYECTADO
# ══════════════════════════════════════════════════════════════

@router.get("/flujo-caja")
def flujo_caja_proyectado(meses: int = 3, u=Depends(get_current_user)):
    """Projected cash flow for the next N months."""
    from datetime import date, timedelta
    import calendar

    resultado = []
    today = date.today()

    for i in range(meses):
        m = today.month + i
        y = today.year
        while m > 12: m -= 12; y += 1
        last_day = calendar.monthrange(y, m)[1]
        fi = f"{y}-{str(m).zfill(2)}-01"
        ff = f"{y}-{str(m).zfill(2)}-{last_day}"
        mes_label = f"{y}-{str(m).zfill(2)}"

        # Inflows: CXC due this month
        cxc_cobrar = query_one("""
            SELECT COALESCE(SUM(saldo),0) as total FROM fin_cxc
            WHERE saldo > 0 AND fecha_vencimiento BETWEEN %s AND %s
        """, (fi, ff))

        # Outflows: CXP due this month + payroll
        cxp_pagar = query_one("""
            SELECT COALESCE(SUM(saldo),0) as total FROM fin_cxp
            WHERE saldo > 0 AND fecha_vencimiento BETWEEN %s AND %s
        """, (fi, ff))

        nomina = query_one("""
            SELECT COALESCE(SUM(neto_a_pagar + aporte_iess_patronal),0) as total
            FROM nom_roles_pago WHERE estado='APROBADO' AND periodo=%s
        """, (mes_label,))
        costo_nomina = float(nomina['total']) if nomina and nomina['total'] else 0
        # If no payroll for future months, estimate from last month
        if costo_nomina == 0 and i > 0:
            ultimo = query_one("SELECT COALESCE(SUM(neto_a_pagar+aporte_iess_patronal),0) as t FROM nom_roles_pago WHERE estado='APROBADO' AND periodo=(SELECT MAX(periodo) FROM nom_roles_pago WHERE estado='APROBADO')")
            costo_nomina = float(ultimo['t']) if ultimo and ultimo['t'] else 0

        ingresos = float(cxc_cobrar['total'])
        egresos = float(cxp_pagar['total']) + costo_nomina

        resultado.append({
            "mes": mes_label,
            "ingresos_esperados": round(ingresos, 2),
            "cxc_por_cobrar": round(float(cxc_cobrar['total']), 2),
            "egresos_esperados": round(egresos, 2),
            "cxp_por_pagar": round(float(cxp_pagar['total']), 2),
            "nomina_estimada": round(costo_nomina, 2),
            "flujo_neto": round(ingresos - egresos, 2),
        })

    return {
        "titulo": "Flujo de Caja Proyectado",
        "meses": resultado,
        "total_ingresos": round(sum(m['ingresos_esperados'] for m in resultado), 2),
        "total_egresos": round(sum(m['egresos_esperados'] for m in resultado), 2),
        "flujo_neto_total": round(sum(m['flujo_neto'] for m in resultado), 2),
    }


# ══════════════════════════════════════════════════════════════
#  19. DASHBOARD FINANCIERO
# ══════════════════════════════════════════════════════════════

@router.get("/dashboard-financiero")
def dashboard_financiero(u=Depends(get_current_user)):
    """Complete financial dashboard."""
    cxc = query_one("SELECT COALESCE(SUM(saldo),0) as total, COUNT(CASE WHEN saldo>0 THEN 1 END) as cuentas FROM fin_cxc")
    cxp = query_one("SELECT COALESCE(SUM(saldo),0) as total, COUNT(CASE WHEN saldo>0 THEN 1 END) as cuentas FROM fin_cxp")

    cxc_vencida = query_one("SELECT COALESCE(SUM(saldo),0) as total FROM fin_cxc WHERE saldo>0 AND fecha_vencimiento<CURRENT_DATE")
    cxp_vencida = query_one("SELECT COALESCE(SUM(saldo),0) as total FROM fin_cxp WHERE saldo>0 AND fecha_vencimiento<CURRENT_DATE")

    bancos = query_one("SELECT COALESCE(SUM(saldo_inicial),0) as total FROM fin_cuentas_bancarias WHERE activa=true")

    ventas_mes = query_one("SELECT COALESCE(SUM(total),0) as t FROM ven_facturas WHERE estado='EMITIDA' AND fecha_emision>=DATE_TRUNC('month',CURRENT_DATE)")
    compras_mes = query_one("SELECT COALESCE(SUM(total),0) as t FROM com_compras WHERE estado='CONFIRMADA' AND fecha>=DATE_TRUNC('month',CURRENT_DATE)")

    # Ratios
    liquidez = round(float(cxc['total']) / float(cxp['total']), 2) if float(cxp['total']) > 0 else 0

    return {
        "cxc": {"total": float(cxc['total']), "cuentas": cxc['cuentas'], "vencida": float(cxc_vencida['total'])},
        "cxp": {"total": float(cxp['total']), "cuentas": cxp['cuentas'], "vencida": float(cxp_vencida['total'])},
        "bancos_saldo": float(bancos['total']),
        "ventas_mes": float(ventas_mes['t']),
        "compras_mes": float(compras_mes['t']),
        "ratio_liquidez": liquidez,
        "capital_trabajo": round(float(cxc['total']) - float(cxp['total']), 2),
    }


@router.get("/ats/xml")
def descargar_ats_xml(mes: str, u=Depends(get_current_user)):
    """Generate ATS XML file for SRI upload."""
    import xml.etree.ElementTree as ET
    import calendar as cal_mod

    data = generar_ats(mes, u)
    empresa = query_one("SELECT * FROM sys_empresas WHERE activa=true LIMIT 1")

    year, month = mes.split('-')
    ruc_empresa = empresa.get("ruc", "") if empresa else ""

    root = ET.Element("iva")
    # Informante
    ET.SubElement(root, "TipoIDInformante").text = "R"
    ET.SubElement(root, "IdInformante").text = ruc_empresa
    ET.SubElement(root, "razonSocial").text = (empresa or {}).get("razon_social", "")
    ET.SubElement(root, "Anio").text = year
    ET.SubElement(root, "Mes").text = month.zfill(2)

    # Map tipo_identificacion to SRI codes
    def tipo_id_sri(tipo):
        mapping = {
            'RUC': '01', 'CEDULA': '02', 'PASAPORTE': '03',
            'CONSUMIDOR_FINAL': '07', 'EXTERIOR': '08',
        }
        return mapping.get((tipo or '').upper(), '02')

    # --- Compras ---
    if data["compras"]:
        compras_el = ET.SubElement(root, "compras")
        for c in data["compras"]:
            det = ET.SubElement(compras_el, "detalleCompras")
            ET.SubElement(det, "codSustento").text = "01"
            ET.SubElement(det, "tpIdProv").text = tipo_id_sri(c.get("tipo_identificacion"))
            ET.SubElement(det, "idProv").text = c.get("identificacion", "")
            ET.SubElement(det, "tipoComprobante").text = "01"
            ET.SubElement(det, "fechaRegistro").text = str(c.get("fecha", ""))[:10].replace("-", "/") if c.get("fecha") else ""
            ET.SubElement(det, "establecimiento").text = str(c.get("num_documento", ""))[:3] if c.get("num_documento") else "001"
            ET.SubElement(det, "puntoEmision").text = str(c.get("num_documento", ""))[4:7] if c.get("num_documento") and len(str(c.get("num_documento","")))>6 else "001"
            ET.SubElement(det, "secuencial").text = str(c.get("num_documento", ""))[-9:] if c.get("num_documento") else "000000001"
            ET.SubElement(det, "baseNoGraIva").text = str(round(float(c.get("subtotal_0", 0)), 2))
            ET.SubElement(det, "baseImponible").text = str(round(float(c.get("subtotal_iva", 0)), 2))
            ET.SubElement(det, "montoIva").text = str(round(float(c.get("iva", 0)), 2))

    # --- Ventas ---
    if data["ventas"]:
        ventas_el = ET.SubElement(root, "ventas")
        for v in data["ventas"]:
            det = ET.SubElement(ventas_el, "detalleVentas")
            ET.SubElement(det, "tpIdCliente").text = tipo_id_sri(v.get("tipo_identificacion"))
            ET.SubElement(det, "idCliente").text = v.get("identificacion", "")
            ET.SubElement(det, "tipoComprobante").text = "18"
            num = v.get("numero_factura", "") or ""
            ET.SubElement(det, "tipoEmision").text = "E"
            ET.SubElement(det, "numeroComprobantes").text = "1"
            ET.SubElement(det, "baseNoGraIva").text = str(round(float(v.get("subtotal_0", 0)), 2))
            ET.SubElement(det, "baseImponible").text = str(round(float(v.get("subtotal_iva", 0)), 2))
            ET.SubElement(det, "montoIva").text = str(round(float(v.get("iva", 0)), 2))

    # --- Ventas por establecimiento ---
    ventas_est = ET.SubElement(root, "ventasEstablecimiento")
    det_est = ET.SubElement(ventas_est, "ventaEst")
    ET.SubElement(det_est, "codEstab").text = "001"
    ET.SubElement(det_est, "ventasEstab").text = str(round(data["resumen"]["total_ventas"], 2))

    # --- Anulados ---
    if data["anulados"]:
        anulados_el = ET.SubElement(root, "anulados")
        for a in data["anulados"]:
            det = ET.SubElement(anulados_el, "detalleAnulados")
            ET.SubElement(det, "tipoComprobante").text = a.get("tipo_doc", "01")
            num = a.get("numero", "") or ""
            ET.SubElement(det, "establecimiento").text = num[:3] if len(num) >= 3 else "001"
            ET.SubElement(det, "puntoEmision").text = num[4:7] if len(num) >= 7 else "001"
            ET.SubElement(det, "secuencialInicio").text = num[-9:] if len(num) >= 9 else num
            ET.SubElement(det, "secuencialFin").text = num[-9:] if len(num) >= 9 else num
            ET.SubElement(det, "autorizacion").text = ""

    xml_str = ET.tostring(root, encoding="unicode", xml_declaration=False)
    xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + xml_str

    return Response(
        content=xml_str.encode("utf-8"),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="ATS_{mes}.xml"'}
    )
