"""
Importación masiva de datos desde Excel/CSV.
Soporta: productos, clientes, proveedores, empleados, plan de cuentas.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from database import query, query_one, execute, insert
from auth import get_current_user
from typing import Optional
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/api/importar", tags=["Importar"])


# ── Helpers ──────────────────────────────────────────────────
def _leer_excel(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(400, "El archivo está vacío")

    # Detectar fila de cabeceras: primera fila con 2+ celdas no vacías
    # (la fila de instrucción de nuestra plantilla tiene solo 1 celda merged)
    header_idx = 0
    for idx, fila in enumerate(filas):
        no_vacias = [c for c in fila if c is not None and str(c).strip() != ""]
        if len(no_vacias) >= 2:
            header_idx = idx
            break

    # Limpiar headers: quitar " *" de campos obligatorios y normalizar
    def _clean_header(h):
        if not h:
            return ""
        s = str(h).strip().lower()
        s = s.replace(" *", "").replace("*", "").strip()
        return s

    headers = [_clean_header(h) for h in filas[header_idx]]

    # Calcular desde qué fila empiezan los datos reales
    # Si había fila de instrucción (header_idx > 0), la siguiente fila es descripción → saltarla
    data_start = header_idx + 1
    if header_idx > 0:
        data_start += 1  # saltar fila de descripciones de la plantilla

    datos = []
    for fila in filas[data_start:]:
        if all(v is None or str(v).strip() == "" for v in fila):
            continue
        datos.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in fila])))
    return headers, datos

def _val(row, *keys, default=""):
    for k in keys:
        v = row.get(k.lower(), "")
        if v and v != "None":
            return v
    return default

def _float(v, default=0.0):
    try: return float(str(v).replace(",", "."))
    except: return default

def _int(v, default=0):
    try: return int(float(str(v)))
    except: return default

def _bool(v):
    return str(v).lower() in ("1", "true", "si", "sí", "yes", "verdadero", "activo")

def _excel_template(titulo: str, columnas: list, ejemplos: list):
    """Genera un Excel con cabecera estilizada y filas de ejemplo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo

    # Estilo cabecera
    hdr_fill = PatternFill("solid", fgColor="1D4ED8")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila de instrucción
    ws.merge_cells(f"A1:{chr(64+len(columnas))}1")
    c = ws.cell(1, 1, f"📋 Plantilla de importación: {titulo}  |  No borrar esta fila ni las cabeceras")
    c.fill = PatternFill("solid", fgColor="0F172A")
    c.font = Font(bold=True, color="94A3B8", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Cabeceras
    for i, (col_id, col_label, col_desc, col_req) in enumerate(columnas, 1):
        c = ws.cell(2, i, col_label + (" *" if col_req else ""))
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = hdr_align
        c.border = border
        ws.column_dimensions[chr(64+i)].width = max(18, len(col_label)+4)

    # Fila de descripción (gris)
    desc_fill = PatternFill("solid", fgColor="1E293B")
    for i, (col_id, col_label, col_desc, col_req) in enumerate(columnas, 1):
        c = ws.cell(3, i, col_desc)
        c.fill = desc_fill
        c.font = Font(color="64748B", size=9, italic=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[3].height = 30

    # Filas de ejemplo
    ex_fills = [PatternFill("solid", fgColor="0F172A"), PatternFill("solid", fgColor="111827")]
    for r, ej in enumerate(ejemplos, 4):
        for i, (col_id, _, _, _) in enumerate(columnas, 1):
            c = ws.cell(r, i, ej.get(col_id, ""))
            c.fill = ex_fills[r % 2]
            c.font = Font(color="E2E8F0", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
#  PRODUCTOS / INVENTARIO
# ══════════════════════════════════════════════════════════════

COLS_PRODUCTOS = [
    ("codigo",      "Código",          "Ej: PROD001",           True),
    ("nombre",      "Nombre",          "Nombre del producto",   True),
    ("descripcion", "Descripción",     "Descripción opcional",  False),
    ("categoria",   "Categoría",       "Ej: Electrónica",       False),
    ("precio_venta","Precio Venta",    "Ej: 25.50",             True),
    ("precio_costo","Precio Costo",    "Ej: 15.00",             False),
    ("stock",       "Stock Inicial",   "Cantidad inicial",      False),
    ("stock_minimo","Stock Mínimo",    "Para alertas de stock", False),
    ("unidad",      "Unidad",          "Ej: UNIDAD, KG, LT",    False),
    ("iva",         "IVA %",           "0, 5 o 15",             False),
    ("activo",      "Activo",          "1=Sí, 0=No",            False),
]

EJEMPLOS_PRODUCTOS = [
    {"codigo":"PROD001","nombre":"Laptop HP 15","descripcion":"Laptop 8GB RAM","categoria":"Tecnología",
     "precio_venta":"850.00","precio_costo":"620.00","stock":"10","stock_minimo":"2","unidad":"UNIDAD","iva":"15","activo":"1"},
    {"codigo":"PROD002","nombre":"Mouse Inalámbrico","descripcion":"Mouse USB","categoria":"Accesorios",
     "precio_venta":"18.50","precio_costo":"9.00","stock":"50","stock_minimo":"5","unidad":"UNIDAD","iva":"15","activo":"1"},
]

@router.get("/plantilla/productos")
def plantilla_productos(u=Depends(get_current_user)):
    buf = _excel_template("Productos", COLS_PRODUCTOS, EJEMPLOS_PRODUCTOS)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"})

@router.post("/productos")
async def importar_productos(file: UploadFile = File(...), u=Depends(get_current_user)):
    headers, datos = _leer_excel(await file.read())
    ok, errores = 0, []
    for i, row in enumerate(datos, 4):
        try:
            codigo = _val(row, "código", "codigo", "code", "cod")
            # inv_productos usa 'descripcion' como nombre del producto
            nombre = _val(row, "nombre", "name", "producto", "descripción", "descripcion", "description")
            if not codigo or not nombre:
                errores.append(f"Fila {i}: código y nombre son obligatorios"); continue

            precio = _float(_val(row, "precio venta", "precio_venta", "precio", "pvp", "price"))
            stock  = _float(_val(row, "stock inicial", "stock", "cantidad"))
            iva    = _float(_val(row, "iva %", "iva", "impuesto"), 15.0)
            activo = _bool(_val(row, "activo", "active", "estado") or "1")

            existe = query_one("SELECT id FROM inv_productos WHERE codigo=%s", (codigo,))
            if existe:
                execute("UPDATE inv_productos SET descripcion=%s, iva_porcentaje=%s, activo=%s WHERE id=%s",
                        (nombre, iva, activo, existe["id"]))
                pid = existe["id"]
            else:
                pid = insert("""
                    INSERT INTO inv_productos (codigo, descripcion, iva_porcentaje, activo, clase)
                    VALUES (%s,%s,%s,%s,'MERCADERIA')
                """, (codigo, nombre, iva, activo))
                # Stock inicial en bodega principal
                if stock > 0:
                    bod = query_one("SELECT id FROM inv_bodegas WHERE es_principal=true AND activa=true LIMIT 1")
                    if bod:
                        execute("INSERT INTO inv_stock (producto_id, bodega_id, cantidad) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                                (pid, bod["id"], stock))
                        execute("UPDATE inv_stock SET cantidad=%s WHERE producto_id=%s AND bodega_id=%s",
                                (stock, pid, bod["id"]))

            # Precio de venta en tabla inv_precios
            if precio > 0:
                execute("INSERT INTO inv_precios (producto_id, tipo_precio_id, precio, activo) VALUES (%s,1,%s,true) ON CONFLICT DO NOTHING",
                        (pid, precio))
                execute("UPDATE inv_precios SET precio=%s WHERE producto_id=%s AND tipo_precio_id=1",
                        (precio, pid))
            ok += 1
        except Exception as e:
            errores.append(f"Fila {i}: {str(e)[:80]}")
    return {"importados": ok, "errores": errores,
            "msg": f"{ok} productos importados. {len(errores)} errores."}


# ══════════════════════════════════════════════════════════════
#  CLIENTES
# ══════════════════════════════════════════════════════════════

COLS_CLIENTES = [
    ("identificacion","RUC / Cédula",     "10 o 13 dígitos",       True),
    ("tipo_id",       "Tipo ID",           "CEDULA, RUC o PASAPORTE",True),
    ("nombre",        "Razón Social",      "Nombre o empresa",      True),
    ("email",         "Email",             "correo@ejemplo.com",    False),
    ("telefono",      "Teléfono",          "Ej: 0991234567",        False),
    ("direccion",     "Dirección",         "Dirección completa",    False),
    ("ciudad",        "Ciudad",            "Ej: Quito",             False),
    ("activo",        "Activo",            "1=Sí, 0=No",            False),
]

EJEMPLOS_CLIENTES = [
    {"identificacion":"1712345678","tipo_id":"CEDULA","nombre":"Juan Pérez","email":"juan@email.com",
     "telefono":"0991234567","direccion":"Av. Principal 123","ciudad":"Quito","activo":"1"},
    {"identificacion":"0912345678001","tipo_id":"RUC","nombre":"Empresa ABC S.A.","email":"info@empresa.com",
     "telefono":"042345678","direccion":"Calle 10 Norte","ciudad":"Guayaquil","activo":"1"},
]

@router.get("/plantilla/clientes")
def plantilla_clientes(u=Depends(get_current_user)):
    buf = _excel_template("Clientes", COLS_CLIENTES, EJEMPLOS_CLIENTES)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_clientes.xlsx"})

@router.post("/clientes")
async def importar_clientes(file: UploadFile = File(...), u=Depends(get_current_user)):
    headers, datos = _leer_excel(await file.read())
    ok, errores = 0, []
    for i, row in enumerate(datos, 4):
        try:
            ident  = _val(row, "ruc / cédula", "ruc/cédula", "ruc/cedula", "cedula", "ruc",
                          "identificacion", "identificación", "id", "ruc / cedula")
            nombre = _val(row, "razón social", "razon social", "razon_social", "nombre",
                          "name", "empresa", "cliente")
            if not ident or not nombre:
                errores.append(f"Fila {i}: identificación y nombre son obligatorios"); continue
            tipo   = _val(row, "tipo id", "tipo_id", "tipo id", "tipo_identificacion", "tipo") \
                     or ("RUC" if len(ident) == 13 else "CEDULA")
            email  = _val(row, "email", "correo")
            tel    = _val(row, "teléfono", "telefono", "phone")
            dir_   = _val(row, "dirección", "direccion", "address")
            ciudad = _val(row, "ciudad", "city")
            activo = _bool(_val(row, "activo", "active", "estado") or "1")

            existe = query_one("SELECT id FROM ven_clientes WHERE identificacion=%s", (ident,))
            if existe:
                execute("""UPDATE ven_clientes SET razon_social=%s, email=%s, telefono=%s,
                    direccion=%s, ciudad=%s, activo=%s WHERE id=%s""",
                    (nombre, email, tel, dir_, ciudad, activo, existe["id"]))
            else:
                insert("""INSERT INTO ven_clientes
                    (identificacion, tipo_identificacion, razon_social, email, telefono, direccion, ciudad, activo)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (ident, tipo, nombre, email, tel, dir_, ciudad, activo))
            ok += 1
        except Exception as e:
            errores.append(f"Fila {i}: {str(e)[:80]}")
    return {"importados": ok, "errores": errores,
            "msg": f"{ok} clientes importados. {len(errores)} errores."}


# ══════════════════════════════════════════════════════════════
#  PROVEEDORES
# ══════════════════════════════════════════════════════════════

COLS_PROVEEDORES = [
    ("ruc",       "RUC",           "RUC del proveedor",     True),
    ("nombre",    "Razón Social",  "Nombre o empresa",      True),
    ("email",     "Email",         "correo@ejemplo.com",    False),
    ("telefono",  "Teléfono",      "Ej: 0991234567",        False),
    ("direccion", "Dirección",     "Dirección completa",    False),
    ("ciudad",    "Ciudad",        "Ej: Quito",             False),
    ("contacto",  "Contacto",      "Nombre del contacto",   False),
    ("activo",    "Activo",        "1=Sí, 0=No",            False),
]

EJEMPLOS_PROVEEDORES = [
    {"ruc":"1790123456001","nombre":"Proveedor XYZ S.A.","email":"ventas@xyz.com",
     "telefono":"022345678","direccion":"Parque Industrial","ciudad":"Quito","contacto":"Carlos Ruiz","activo":"1"},
]

@router.get("/plantilla/proveedores")
def plantilla_proveedores(u=Depends(get_current_user)):
    buf = _excel_template("Proveedores", COLS_PROVEEDORES, EJEMPLOS_PROVEEDORES)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_proveedores.xlsx"})

@router.post("/proveedores")
async def importar_proveedores(file: UploadFile = File(...), u=Depends(get_current_user)):
    headers, datos = _leer_excel(await file.read())
    ok, errores = 0, []
    for i, row in enumerate(datos, 4):
        try:
            ruc    = _val(row, "ruc", "identificacion", "identificación", "id")
            nombre = _val(row, "razón social", "razon social", "razon_social", "nombre", "name")
            if not ruc or not nombre:
                errores.append(f"Fila {i}: RUC y nombre son obligatorios"); continue
            email    = _val(row, "email", "correo")
            tel      = _val(row, "teléfono", "telefono", "phone")
            dir_     = _val(row, "dirección", "direccion")
            ciudad   = _val(row, "ciudad", "city")
            contacto = _val(row, "contacto", "contact", "contacto_nombre")
            activo   = _bool(_val(row, "activo") or "1")
            tipo     = "RUC" if len(ruc) == 13 else "CEDULA"

            existe = query_one("SELECT id FROM com_proveedores WHERE identificacion=%s", (ruc,))
            if existe:
                execute("""UPDATE com_proveedores SET razon_social=%s, email=%s, telefono=%s,
                    direccion=%s, ciudad=%s, contacto_nombre=%s, activo=%s WHERE id=%s""",
                    (nombre, email, tel, dir_, ciudad, contacto, activo, existe["id"]))
            else:
                insert("""INSERT INTO com_proveedores
                    (identificacion, tipo_identificacion, razon_social, email, telefono,
                     direccion, ciudad, contacto_nombre, activo)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (ruc, tipo, nombre, email, tel, dir_, ciudad, contacto, activo))
            ok += 1
        except Exception as e:
            errores.append(f"Fila {i}: {str(e)[:80]}")
    return {"importados": ok, "errores": errores,
            "msg": f"{ok} proveedores importados. {len(errores)} errores."}


# ══════════════════════════════════════════════════════════════
#  EMPLEADOS
# ══════════════════════════════════════════════════════════════

COLS_EMPLEADOS = [
    ("cedula",        "Cédula",         "10 dígitos",            True),
    ("nombres",       "Nombres",        "Primer y segundo nombre",True),
    ("apellidos",     "Apellidos",      "Apellidos completos",   True),
    ("email",         "Email",          "correo@empresa.com",    False),
    ("telefono",      "Teléfono",       "Ej: 0991234567",        False),
    ("cargo",         "Cargo",          "Ej: Vendedor",          False),
    ("departamento",  "Departamento",   "Ej: Ventas",            False),
    ("salario_base",  "Salario Base",   "Ej: 475.00",            True),
    ("fecha_ingreso", "Fecha Ingreso",  "YYYY-MM-DD",            False),
    ("tipo_contrato", "Tipo Contrato",  "INDEFINIDO,FIJO,PRUEBA",False),
    ("region",        "Región",         "SIERRA o COSTA",        False),
]

EJEMPLOS_EMPLEADOS = [
    {"cedula":"1712345678","nombres":"Juan Carlos","apellidos":"Pérez López","email":"juan@empresa.com",
     "telefono":"0991234567","cargo":"Vendedor","departamento":"Ventas",
     "salario_base":"475.00","fecha_ingreso":"2024-01-15","tipo_contrato":"INDEFINIDO","region":"SIERRA"},
]

@router.get("/plantilla/empleados")
def plantilla_empleados(u=Depends(get_current_user)):
    buf = _excel_template("Empleados", COLS_EMPLEADOS, EJEMPLOS_EMPLEADOS)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_empleados.xlsx"})

@router.post("/empleados")
async def importar_empleados(file: UploadFile = File(...), u=Depends(get_current_user)):
    headers, datos = _leer_excel(await file.read())
    ok, errores = 0, []
    for i, row in enumerate(datos, 4):
        try:
            cedula   = _val(row, "cédula", "cedula", "id", "dni")
            nombres  = _val(row, "nombres", "nombre", "name")
            apellidos = _val(row, "apellidos", "apellido", "lastname")
            salario  = _float(_val(row, "salario base", "salario_base", "salario", "sueldo"))
            if not cedula or not nombres or not apellidos:
                errores.append(f"Fila {i}: cédula, nombres y apellidos son obligatorios"); continue
            if salario <= 0:
                errores.append(f"Fila {i}: salario inválido"); continue

            email    = _val(row, "email", "correo")
            tel      = _val(row, "teléfono", "telefono")
            cargo    = _val(row, "cargo", "puesto", "position")
            depto    = _val(row, "departamento", "department")
            from datetime import date as _date
            fi_raw   = _val(row, "fecha ingreso", "fecha_ingreso")
            fi       = fi_raw if fi_raw else str(_date.today())  # hoy como fallback
            contrato = _val(row, "tipo contrato", "tipo_contrato") or "INDEFINIDO"
            region   = _val(row, "región", "region") or "SIERRA"

            existe = query_one("SELECT id FROM nom_empleados WHERE cedula=%s", (cedula,))
            if existe:
                execute("""UPDATE nom_empleados SET nombres=%s, apellidos=%s, email=%s,
                    telefono=%s, cargo=%s, departamento=%s, salario_base=%s,
                    fecha_ingreso=%s, tipo_contrato=%s, region=%s WHERE id=%s""",
                    (nombres, apellidos, email, tel, cargo, depto, salario,
                     fi, contrato, region, existe["id"]))
            else:
                insert("""INSERT INTO nom_empleados
                    (cedula, nombres, apellidos, email, telefono, cargo, departamento,
                     salario_base, fecha_ingreso, tipo_contrato, region, activo)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,true)""",
                    (cedula, nombres, apellidos, email, tel, cargo, depto,
                     salario, fi, contrato, region))
            ok += 1
        except Exception as e:
            errores.append(f"Fila {i}: {str(e)[:80]}")

    return {"importados": ok, "errores": errores,
            "msg": f"{ok} empleados importados. {len(errores)} errores."}
